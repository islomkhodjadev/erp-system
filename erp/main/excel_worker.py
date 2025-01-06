
import math
from decimal import Decimal, InvalidOperation, ROUND_UP

def safe_decimal(value):
    """
    Safely convert input to a Decimal, replacing NaN, None, or invalid values with 0.0.
    """
    if value is None:
        return Decimal('0.0')
    
    # If value is a string and it's 'nan', 'none', or empty, treat as 0.0
    if isinstance(value, str) and value.strip().lower() in {'nan', 'none', ''}:
        return Decimal('0.0')

    # Check if the value is an actual NaN (float)
    if isinstance(value, float) and math.isnan(value):
        return Decimal('0.0')

    try:
        # Try converting the value to Decimal
        return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_UP)
    except (InvalidOperation, TypeError, ValueError):
        # In case of invalid value, return 0.0
        return Decimal('0.0')

# Now applying the safe_decimal function to handle 'NaN' or invalid values.


from dataclasses import dataclass
from decimal import Decimal
from datetime import date
import pandas as pd
import openpyxl
from dataclasses import dataclass  # Correct import
from typing import List

# Define data classes
@dataclass
class ProfileData:
    id_user: str
    telegram_id: str
    telegram_username: str
    password: str
    language: str
    is_loggined: bool
    is_blocked: bool

@dataclass
class DebtData:
    profile_id_user: str
    total_borrowed: Decimal
    total_paid: Decimal
    remaining_balance: Decimal

@dataclass
class DebtMovementData:
    debt_profile_id_user: str
    movement_type: str
    amount: Decimal
    movement_date: date

@dataclass
class UserData:
    name: str
    profile: ProfileData
    debt: DebtData
    debt_movement: List[DebtMovementData]

# Function to extract all sheets from an Excel file
def extract_all_sheets_to_dfs(excel_file):
    """
    Extracts all sheets from an Excel file, stores them in a dictionary of DataFrames,
    ignoring the sheet with name "calculator" and protected sheets.
    
    Args:
        excel_file (str): Path to the Excel file.
        
    Returns:
        dict: A dictionary where keys are sheet names and values are DataFrames.
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)  # data_only=True gets calculated values
        sheet_dfs = {}

        # Iterate through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            # Skip the sheet named 'calculator' or any protected sheet
            sheet = workbook[sheet_name]
            if sheet_name.lower() == "calculator" or sheet.protection.sheet:
                continue

            # Read the sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Convert the sheet to a DataFrame and store it in the dictionary
            df = pd.DataFrame(data)
            sheet_dfs[sheet_name] = df

        if sheet_dfs:
            return sheet_dfs  # Return dictionary of DataFrames
        else:
            return {}  # Return empty dictionary if no valid data

    except FileNotFoundError:
        return {}  # Return empty dictionary in case of error
    except Exception as e:
        return {}  # Return empty dictionary in case of error

        
# Function to process sheets and generate object-type output
def process_sheets(sheet_dfs):
    """
    Processes each DataFrame from the sheets dictionary and generates UserData objects.
    
    Args:
        sheet_dfs (dict): Dictionary of DataFrames from all sheets.
        
    Returns:
        List[UserData]: List of UserData objects.
    """
    user_data_list = []

    for sheet_name, combined_df in sheet_dfs.items():
        
        # Extract and clean data
        combined_df['date'] = combined_df[1].astype(str).str.replace(r'[\.\\]', '-', regex=True)
        combined_df['date'] = pd.to_datetime(combined_df['date'], errors='coerce')

        combined_df['Income'] = pd.to_numeric(combined_df[2], errors='coerce')
        combined_df['expense'] = pd.to_numeric(combined_df[3], errors='coerce')
        combined_df['left'] = pd.to_numeric(combined_df[4], errors='coerce')

        # Extract user info from the first two rows
        user_id = str(combined_df.iloc[0, 5])
        password = str(combined_df.iloc[1, 5])

        # Filter rows with valid dates
        valid_data = combined_df[pd.notna(combined_df['date'])]
        
        # Calculate total income, total expense, and final balance based on valid dates
        total_income = valid_data['Income'].sum()
        total_expense = valid_data['expense'].sum()
        final_balance = valid_data['left'].iloc[-1] if not valid_data.empty and 'left' in combined_df.columns else 0
        

        # Generate DebtMovementData
        debt_movements = []
        for _, row in combined_df.iterrows():
            if pd.notna(row['date']):
                if pd.notna(row['Income']):
                    debt_movements.append(
                        DebtMovementData(
                            debt_profile_id_user=user_id,
                            movement_type="debt",
                            amount=Decimal(row['Income']),
                            movement_date=row['date'].date() if not pd.isna(row['date']) else None
                        )
                    )
                if pd.notna(row['expense']):
                    debt_movements.append(
                        DebtMovementData(
                            debt_profile_id_user=user_id,
                            movement_type="paid",
                            amount=Decimal(row['expense']),
                            movement_date=row['date'].date() if not pd.isna(row['date']) else None
                        )
                    )

        # Create ProfileData and DebtData
        profile = ProfileData(

            id_user=user_id,
            # telegram_id=f"",
            # telegram_username=f"",
            password=password,
            language="uz",
            is_loggined=False,
            is_blocked=False
        )

        debt = DebtData(
            profile_id_user=user_id,
            total_borrowed=Decimal(total_income),
            total_paid=Decimal(total_expense),
            remaining_balance=Decimal(final_balance)
        )

        # Create UserData object
        user_data = UserData(
            name = str(sheet_name),
            profile=profile,
            debt=debt,
            debt_movement=debt_movements
        )

        user_data_list.append(user_data)

    return user_data_list

def usage_debt(excel):

    sheet_dfs = extract_all_sheets_to_dfs(excel)
    user_data_list = process_sheets(sheet_dfs)
    return user_data_list
