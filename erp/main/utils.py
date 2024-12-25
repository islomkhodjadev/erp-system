import pandas as pd
import openpyxl


def extract_process_and_combine_sheets(excel_file, sheet_names):
    """
    Extracts specified sheets from an Excel file, processes the data (extracts code, name, packages, and price),
    reshapes the data, and performs price division by packages.

    Args:
        excel_file (str): Path to the Excel file.
        sheet_names (list): List of sheet names to extract.

    Returns:
        pd.DataFrame: Processed and reshaped DataFrame.
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)  # data_only=True gets calculated values
        combined_data = []

        # Extract data from sheets
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue

            # Read the sheet
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Convert the sheet to a DataFrame
            df = pd.DataFrame(data)

            # Append to the combined data
            if not df.empty:
                combined_data.append(df)

        if not combined_data:
            return pd.DataFrame()  # Return an empty DataFrame if no valid data

        # Combine all DataFrames
        combined_df = pd.concat(combined_data, ignore_index=True, axis=0)

        # Extract the first 3, 4, or 5 digits (code) and the words after the digits (name)
        combined_df['code'] = combined_df[2].astype(str).str.extract(r'(\d{3,5})')
        combined_df['name'] = combined_df[2].astype(str).str.extract(r'(\D.*)')  # Extract the words after the number

        # Extract numeric values from columns 5 and 6
        combined_df['num_col_5'] = combined_df[5].astype(str).str.extract(r'(\d+)')
        combined_df['num_col_6'] = combined_df[6].astype(str).str.extract(r'(\d+)')

        # Combine numeric values from columns 5 and 6 into a new 'packages' column
        combined_df['packages'] = combined_df['num_col_5'].fillna('') + ' ' + combined_df['num_col_6'].fillna('')
        combined_df['packages'] = combined_df['packages'].str.strip()  # Remove extra spaces

        # Extract or assign the 'price' from column 7
        combined_df['price'] = combined_df[7].astype(str).str.extract(r'(\d+(\.\d+)?)')[0]  # Extract the first capturing group

        # Create a new DataFrame with 'code', 'name', 'packages', and 'price' columns
        new_df = combined_df[['code', 'name', 'packages', 'price']]

        # Group by 'code' and reshape
        grouped = new_df.groupby('code')

        # Create a list to store the reshaped rows
        reshaped_data = []

        for code, group in grouped:
            row = {"code": code}
            for idx, (_, row_data) in enumerate(group.iterrows(), start=2):  # Start index from 2
                if idx > 2:  # Skip writing the first row as `name_1`, `packages_1`, `price_1`
                    row[f"name_{idx-2}"] = row_data['name']
                    row[f"packages_{idx-2}"] = row_data['packages']
                    row[f"price_{idx-2}"] = row_data['price']  # Add the price for each row
            reshaped_data.append(row)

        # Convert the reshaped data into a new DataFrame
        final_df = pd.DataFrame(reshaped_data)

        # Loop through all columns starting with 'price_' and 'packages_' to perform the division
        for i in range(1, 7):  # Loop from 1 to 6 to handle price_1, ..., price_6
            price_column = f'price_{i}'
            packages_column = f'packages_{i}'

            # Ensure both price and packages columns are numeric for division
            final_df[price_column] = pd.to_numeric(final_df[price_column], errors='coerce')
            final_df[packages_column] = pd.to_numeric(final_df[packages_column], errors='coerce')

            # Perform the division and update the price column
            final_df[price_column] = final_df[price_column] / final_df[packages_column]

        # Return the processed DataFrame
        return final_df

    except FileNotFoundError:
        return pd.DataFrame()  # Return an empty DataFrame in case of error
    except Exception as e:
        return pd.DataFrame()  # Return an empty DataFrame in case of error

















import os
import django
import sys
from dotenv import load_dotenv
import requests

# Load environment variables from .env file
load_dotenv()
from .models import Profile, Chats
# Use your bot token from .env
API_TOKEN = os.getenv("tg_token")



def inform_order_user_gorup(user_id, message):
    """
    Function to send a message to both a user and a group chat (like an 'order' chat) via the Telegram Bot API.
    
    Args:
    - user_id (str): The Telegram user ID.
    - message (str): The message to be sent.
    - group_chat_id (str): The ID of the group chat (e.g., 'order' chat).
    
    Returns:
    - str: Confirmation or error message.
    """
    try:

        # Fetch the user's username from the Profile model (assuming Profile model exists)
        group_chat_id = Chats.get_chat_id_by_type("order")

        profile = Profile.objects.get(telegram_id=user_id)  # Fetch profile based on the user_id
        user_username = profile.telegram_username if profile.telegram_username else "User"

        # Send message to the individual user
        send_message_to_telegram(user_id, message)
        
        group_message = f"Новый заказ от @{user_username} {profile.name if profile.name else ''}\nYangi buyurtma @{user_username} {profile.name if profile.name else ''}:\n\n{message}"


        # Send message to the group chat
        send_message_to_telegram(group_chat_id, group_message)

        return "Message sent to both user and group chat."
    
    except Profile.DoesNotExist:
        return "User profile not found."
    except Exception as e:
        return f"An error occurred: {str(e)}"


def send_message_to_telegram(chat_id, message):
    
    """
    Helper function to send a message via Telegram Bot API.
    
    Args:
    - chat_id (str): The Telegram ID of the user or group chat.
    - message (str): The message to be sent.
    """
    url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
    data = {
        'chat_id': chat_id,
        'text': message,
        "parse_mode": "HTML"
    }

    # Send the message via Telegram Bot API
    response = requests.post(url, data=data)

    if response.status_code != 200:
        print(f"Failed to send message. Status code: {response.status_code}")
    else:
        print(f"Message sent to chat ID {chat_id}")









from dataclasses import dataclass
from decimal import Decimal
from datetime import date
from typing import List


@dataclass
class ProfileData:
    id_user: str
    telegram_id: str = ""
    telegram_username: str = ""
    password: str = ""
    language: str = "uz"
    is_loggined: bool = False
    is_blocked: bool = False


@dataclass
class DebtData:
    profile_id_user: str
    total_borrowed: Decimal
    total_paid: Decimal
    remaining_balance: Decimal


@dataclass
class DebtMovementData:
    debt_profile_id_user: str
    movement_type: str  # 'debt' or 'paid'
    amount: Decimal
    movement_date: date


# Data structure for a single profile and debt, but many movements
@dataclass
class UserData:
    profile: ProfileData
    debt: DebtData
    debt_movement: List[DebtMovementData]





import pandas as pd
import openpyxl
from datetime import datetime
from decimal import Decimal

# Function to extract all sheets from an Excel file
def extract_all_sheets_to_dfs(excel_file):
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_dfs = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet_name.lower() == "calculator" or sheet.protection.sheet:
                continue

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            df = pd.DataFrame(data)
            sheet_dfs[sheet_name] = df

        return sheet_dfs
    except Exception as e:
        return {}
def process_sheets_to_userdata(sheet_dfs):
    user_data_list = []

    for sheet_name, combined_df in sheet_dfs.items():

        sheet_name = sheet_name.encode('utf-8').decode('utf-8')  # Decode the name in case it has non-ASCII characters
        
        # Clean and convert the 'date' column directly to datetime, then format it
        combined_df['date'] = pd.to_datetime(combined_df[1].astype(str), format='%Y/%m/%d', errors='coerce').dt.strftime('%Y/%m/%d')

        # Drop rows where 'date' is NaT (Invalid date)
        combined_df = combined_df.dropna(subset=['date']).reset_index(drop=True)

        # Extract other columns (Income, Expense, Left, User ID, and Password)
        combined_df['Income'] = combined_df[2].astype(str).str.extract(r'(\d+\.\d+|\d+)')
        combined_df['expense'] = combined_df[3].astype(str).str.extract(r'(\d+\.\d+|\d+)')
        combined_df['left'] = combined_df[4].astype(str).str.extract(r'(\d+\.\d+|\d+)')

        # Extract user_id and password from the first two rows
        combined_df['user_id'] = str(combined_df.iloc[0, 5])
        combined_df['password'] = str(combined_df.iloc[1, 5])
        if combined_df.empty:
            continue  # Skip this sheet if it's empty

        # Create a new DataFrame with the relevant columns
        new_df = combined_df[['date', 'Income', 'expense', 'left', 'user_id', 'password']]
        new_df = new_df.iloc[4:].reset_index(drop=True)  # Remove the first 4 rows

        # Convert columns to numeric, handling errors and applying rounding
        new_df['Income'] = pd.to_numeric(new_df['Income'], errors='coerce').apply(lambda x: round(x, 2) if pd.notna(x) else x)
        new_df['expense'] = pd.to_numeric(new_df['expense'], errors='coerce').apply(lambda x: round(x, 2) if pd.notna(x) else x)
        new_df['left'] = pd.to_numeric(new_df['left'], errors='coerce').apply(lambda x: round(x, 2) if pd.notna(x) else x)

        # Extract user info from the first row
        user_id = new_df.iloc[0]['user_id']
        password = new_df.iloc[0]['password']

        # Calculate totals and final balance
        total_income = new_df['Income'].sum()
        total_expense = new_df['expense'].sum()
        final_balance = total_expense - total_income

        # Create DebtMovements list
        debt_movements = []
        remaining_balance = final_balance

        for _, row in new_df.iterrows():
            if pd.notna(row['date']):
                if pd.notna(row['Income']):
                    debt_movements.append(DebtMovementData(
                        debt_profile_id_user=user_id,
                        movement_type="debt",
                        amount=Decimal(str(row['Income'])),
                        movement_date=datetime.strptime(row['date'], '%Y/%m/%d').date()
                    ))
                    remaining_balance += row['Income']

                if pd.notna(row['expense']):
                    debt_movements.append(DebtMovementData(
                        debt_profile_id_user=user_id,
                        movement_type="paid",
                        amount=Decimal(str(row['expense'])),
                        movement_date=datetime.strptime(row['date'], '%Y/%m/%d').date()
                    ))
                    remaining_balance -= row['expense']

        # Create ProfileData and DebtData
        profile_data = ProfileData(
            id_user=user_id,
            telegram_id="tg_" + user_id,
            telegram_username=user_id + "_username",
            password=password,
            language="uz",  # Assuming "uz" for now
            is_loggined=True,
            is_blocked=False
        )

        debt_data = DebtData(
            profile_id_user=user_id,
            total_borrowed=Decimal(str(total_income)),
            total_paid=Decimal(str(total_expense)),
            remaining_balance=Decimal(str(final_balance))
        )

        user_data = UserData(
            profile=profile_data,
            debt=debt_data,
            debt_movement=debt_movements
        )

        user_data_list.append(user_data)

    return user_data_list

def use_debt_save(excel):
# Example usage:
    sheet_dfs = extract_all_sheets_to_dfs(excel)
    user_data_list = process_sheets_to_userdata(sheet_dfs)
    create_or_update_user_data(user_data_list)

from django.db import transaction
from decimal import Decimal
from datetime import date
from main.models import Profile, Debt, DebtMovement
from decimal import Decimal, InvalidOperation


import math
from decimal import Decimal, InvalidOperation, ROUND_UP

def safe_decimal(value):
    """
    Safely convert input to a Decimal, replacing NaN, None, or invalid values with 0.0.
    """
    if value is None:
        return Decimal('0.0')
    
    # If value is a string and it's 'nan', 'none', or empty, treat as 0.0
    if isinstance(value, str) and value.strip().lower() in {'nan', 'none', ''}:
        return Decimal('0.0')

    # Check if the value is an actual NaN (float)
    if isinstance(value, float) and math.isnan(value):
        return Decimal('0.0')

    try:
        # Try converting the value to Decimal
        return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_UP)
    except (InvalidOperation, TypeError, ValueError):
        # In case of invalid value, return 0.0
        return Decimal('0.0')
    
def create_or_update_user_data(user_data_list):
    """
    Transforms a list of UserData objects into database records for Profile, Debt, and DebtMovement.
    """
    # Start a database transaction to ensure atomicity
    
    for user_data in user_data_list:
        # 1. Create or Update Profile
        profile, created = Profile.objects.get_or_create(
            
            id_user=user_data.profile.id_user,
            defaults={
                "name": user_data.name,
                'telegram_id': user_data.profile.telegram_id,
                'telegram_username': user_data.profile.telegram_username,
                'password': user_data.profile.password,
                'language': user_data.profile.language,
                'is_loggined': user_data.profile.is_loggined,
                'is_blocked': user_data.profile.is_blocked
            })
        
        # 2. Create or Update Debt
        debt, created = Debt.objects.get_or_create(
            profile=profile,
            defaults={
                'total_borrowed': user_data.debt.total_borrowed,
                'total_paid': user_data.debt.total_paid,
                'remaining_balance': user_data.debt.remaining_balance,
            })
        # 3. Create DebtMovements
        for movement in user_data.debt_movement:
            
            DebtMovement.objects.create(
                debt=debt,
                movement_type=movement.movement_type,
                amount=movement.amount,
                movement_date=movement.movement_date)
        
        # After all movements are added, recalculate the debt balance
        debt.save()

    